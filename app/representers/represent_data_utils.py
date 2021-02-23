from __future__ import print_function

import json
import os.path
import os.path as op
import pickle
import re
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.policy import SMTPUTF8

import openpyxl
import pandas as pd
import requests
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from pyexcel import merge_all_to_a_book
from sqlalchemy import or_, JSON, cast, String, literal
from telebot import types
from app.config import MessageStatus, PROJECT_DIR
from app.db import IkeaProduct, session_scope
from app.logging_config import logger
from app.utils import stringify
import typing as t

PRICES = {
    10: 1.5,
    20: 1.45,
    100: 1.40,
    200: 1.35,
    3000: 1.30,
    5000: 1.29,
    10000: 1.27,
    100000: 1.25,
}

PRICES_PL = {
    10: 12,
    20: 11.6,
    50: 11.1,
    100: 10.8,
    200: 10.8,
    500: 10.8,
    1000: 10.5,
    3000: 10.5,
    5000: 10.5,
    10000: 10.5,
}

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# The ID and range of a sample spreadsheet.
SAMPLE_SPREADSHEET_ID = '1OQstaXj8e7wWv4-VM-uArPW0rE32kj_Eak056WnVDXg'
SAMPLE_RANGE_NAME = 'Лист1!A2:A30000'

with open(f"{PROJECT_DIR}/data/search_text.pk", "rb") as f:
    SEARCH_TEXT = pickle.load(f)

parsels_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True,
                                             one_time_keyboard=False,
                                             row_width=2)
# parsels_keyboard.add(
#     types.KeyboardButton(text=MessageStatus.LOAD_FROM_IKEA_2),
#     types.KeyboardButton(text=MessageStatus.LOAD_FROM_IKEA_3),
#     types.KeyboardButton(text=MessageStatus.UPDATE_GOOGLE_SHEETS),
#     types.KeyboardButton(text=MessageStatus.UPDATE_GOOGLE_TREKING),
#     types.KeyboardButton(text=MessageStatus.DIFFERENCE),
#     types.KeyboardButton(text=MessageStatus.UPDATE_UA),
#     types.KeyboardButton(text=MessageStatus.UPDATE_PL),
# )


parsels_keyboard.add(
    types.KeyboardButton(text=MessageStatus.DATA_FROM_IKEA_PL),
    types.KeyboardButton(text=MessageStatus.DATA_FROM_IKEA_UA),
    types.KeyboardButton(text=MessageStatus.FILL_DOC_FIELDS)
)


def get_url_cost(code: str):
    r = requests.get("http://sik.search.blue.cdtapps.com/ua/uk/search-result-page", params={
        "q": code
    })
    res = r.json()['searchResultPage']['productWindow']
    if len(res) == 0:
        return None, None
    return res[0]['pipUrl'], res[0]['priceNumeral']


def get_data(row):
    data = {}
    ru_data = row.ru_data or {}

    ukr_fields = ['Название_позиции', 'Описание', 'Поисковые_запросы']

    data.update(
        {
            key: stringify(value)
            for key, value in (row.pl_data or {}).items()
            if key not in ru_data
        }
    )

    data.update(
        {
            key + ('_укр' if key in ukr_fields else ''): stringify(value)
            for key, value in (row.ua_data or {}).items()
        }
    )

    data.update(
        {
            key: stringify(value)
            for key, value in ru_data.items()
        }
    )
    print('data keys', data.keys())
    print('ua items', row.ua_data.keys())
    return data


def update_delivery():
    """TODO разобраться и наладить
    Пока не поддерживается"""
    import gspread
    import requests
    from oauth2client.service_account import ServiceAccountCredentials
    import pandas as pd

    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
             "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]

    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        '../../data/ikeaparses-google.json', scope)
    client = gspread.authorize(credentials)

    spreadsheet = client.open('Доставка_Ikea')

    r = requests.post("http://testapi.novaposhta.ua/v2.0/en/json/getDocumentList/", data=json.dumps({
        "apiKey": "cdff5758a96e79dc6b5e5a66776ad9fd",
        "modelName": "InternetDocument",
        "calledMethod": "getDocumentList",
    }), headers={
        "Content-Type": "application/json"
    })
    if "data" not in r.json():
        return "Превышено число запросов в минуту"
    req = [{
        "DocumentNumber": d["IntDocNumber"],
        "Phone": d["RecipientsPhone"],
    } for d in r.json()["data"]]

    r = requests.post("http://testapi.novaposhta.ua/v2.0/en/json/getDocumentList/", data=json.dumps({
        "apiKey": "cdff5758a96e79dc6b5e5a66776ad9fd",
        "modelName": "TrackingDocument",
        "calledMethod": "getStatusDocuments",
        "methodProperties": {
            "Documents": req
        }

    }), headers={
        "Content-Type": "application/json"
    })

    doc = pd.DataFrame.from_dict(r.json()["data"])
    doc["Packaging"] = doc["Packaging"].apply(lambda x: "\n".join(c["Description"] for c in x))

    for_drop = ["StatusCode", "LoyaltyCardSender", "MarketplacePartnerToken"]
    doc = doc.drop(columns=[x for x in doc.columns if x in for_drop])
    doc = doc.drop(columns=[x for x in doc.columns if "Ref" in x])
    from app.config import DEV_STATUS_TRANS
    doc = doc.rename(columns=DEV_STATUS_TRANS)
    doc.to_csv("data.csv")

    with open('../../data/data.csv', 'r') as file_obj:
        content = file_obj.read().encode('utf-8')
        client.import_csv(spreadsheet.id, data=content)

    return len(doc)


def update_google_sheets():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('../../data/token.pickle'):
        with open('../../data/token.pickle', 'rb') as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                '../../data/credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('../../data/token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                range=SAMPLE_RANGE_NAME).execute()
    values = result.get('values', [])
    codes = []
    for row in values:
        codes.append(row[0].replace(".", "").strip("0"))
    items = s.query(IkeaProduct.code, IkeaProduct.ru_data).where(
        IkeaProduct.code.in_(codes)).all()
    items = {i.code: i for i in items}
    res = []
    for code in codes:
        if code in items:
            res.append(10 if items[code].avilable else 0)
        else:
            res.append(0)

    service.spreadsheets().values().iter_update(
        spreadsheetId=SAMPLE_SPREADSHEET_ID, range="Лист1!D2:D10000",
        valueInputOption='RAW', body={
            'values': [[r] for r in res],
        }).execute()

    res = []
    for code in codes:
        if code in items:
            pr = float(str(items[code].data["Цена"]).replace(" ", ""))
            for p in PRICES:
                if p > pr:
                    res.append(round(pr * PRICES[p], 2))
                    break
        else:
            res.append(-1)
    service.spreadsheets().values().iter_update(
        spreadsheetId=SAMPLE_SPREADSHEET_ID, range="Лист1!E2:E10000",
        valueInputOption='RAW', body={
            'values': [[r] for r in res],
        }).execute()


import pyexcel


def save_xlsx(lines, filename):
    keys = sum(map(lambda x: list(x.keys()), lines), [])
    default = {
        key: None
        for key in keys
    }
    lines = [
        {
            **default,
            **line
        }
        for line in lines
    ]
    pyexcel.save_as(records=lines, dest_file_name=filename)


def get_tags(row):
    tags = get_data(row).get('tags', [])
    if isinstance(tags, list):
        return tags
    return tags.split(', ')


# TODO Переписать тут все. Создать классы, и в них инкапсулировать логику


def get_groups(items):
    groups = []
    for i in items:
        tags = get_tags(i)
        if len(tags) > 2:
            for t in tags:
                groups.append(t)
    groups = list(set(groups))

    res = []
    done = []

    for i in items:
        tags_orig = []
        for t in get_tags(i):
            if t not in tags_orig and t != "":
                tags_orig.append(t)
        if len(tags_orig) > 2:
            if tags_orig[-2] not in done:
                res.append({
                    # "Номер_группы": groups.index(i.tags[-2]),
                    "Идентификатор_группы": groups.index(tags_orig[-2]),
                    "Название_группы": tags_orig[-2],
                    # "Номер_родителя": groups.index(i.tags[-3]),
                    "Идентификатор_родителя": groups.index(tags_orig[-3]),
                })
                done.append(tags_orig[-2])

            if tags_orig[-3] not in done:
                res.append({
                    # "Номер_группы": groups.index(i.tags[-3]),
                    "Идентификатор_группы": groups.index(tags_orig[-3]),
                    "Название_группы": tags_orig[-3],
                    # "Номер_родителя": "",
                    "Идентификатор_родителя": "",
                })
                done.append(tags_orig[-3])

    return res, groups


def save_ikea_product_to_csv(country, filename):
    def converter(value):
        if isinstance(value, str):
            return value
        if isinstance(value, list):
            return ', '.join(value)
        return str(value)

    products = IkeaProduct.query.filter(IkeaProduct.country == country, IkeaProduct.ru_data.isnot(None)).order_by(
        IkeaProduct.id.desc()).limit(1000).all()
    lines = []
    for row in products:
        line = {}
        try:
            data = row.pl_data or {}
            data.update(row.ua_data or {})
            line.update({
                **data,
                **row.ru_data
            })
            line = {
                str(key): converter(value)
                for key, value in line.items()
            }
            lines.append(line)
        except:
            print(row)

    save_xlsx(lines, filename)


def update_document(xlsx_filepath: str):
    """
    Принимает путь до документа в определенном формате
    Пример файла https://drive.google.com/file/d/1Xtj5JeSyL7MJY9QRrlvBezfA_3PJJaD1/view?usp=sharing
    Функция обновляет колонки Цена и Наличие
    последними данными из бд и перезаписывает файл по тому же пути
    Используются поля таблички
    IkeaProduct: code, country, is_available, data
    """
    assert xlsx_filepath.endswith('.xlsx')

    # Open xlsx file
    workbook = openpyxl.load_workbook(xlsx_filepath)
    products_sheet = workbook['Export Products Sheet']
    first_row = products_sheet[1]
    column_names = {
        cell.value: i for i, cell in enumerate(first_row)
    }
    # Collect ikea product codes
    codes = {}
    for i, row in enumerate(products_sheet.iter_rows()):
        producer_cell = row[column_names['Производитель']]
        code_cell = row[column_names['Код_товара']]
        if producer_cell.value == 'IKEA':
            code = code_cell.value.replace('.', '')
            codes[code] = i

    # Request to DB
    with session_scope() as session:
        query = session.query(IkeaProduct.data, IkeaProduct.code, IkeaProduct.is_available).filter(
            IkeaProduct.code.in_(codes.keys()),
            IkeaProduct.country == 'UA'
        )
        products: t.List[IkeaProduct] = query.all()

        def set_cell_value(code, col_name, value):
            products_sheet.cell(column=column_names[col_name] + 1,
                                row=codes[code] + 1).value = value

        print('Products collected', len(products))
        print('Codes collected', len(codes))
        for product in products:
            set_cell_value(product.code, 'Наличие', '+' if product.is_available else '-')
            data = product.data or {}
            price = float(str(data.get('price', '0') or '0').replace(',', '.'))
            if not price:
                continue
            for p in PRICES:
                if p > price:
                    price *= PRICES[p]
                    break
            set_cell_value(product.code, 'Цена', price)
    workbook.save(xlsx_filepath)


if __name__ == '__main__':
    update_document('you_can_delete_this.xlsx')


def get_ua_items(ignore_codes=[], mode="UA"):
    items = IkeaProduct.query.filter(IkeaProduct.country == mode, IkeaProduct.is_failed == False,
                                     IkeaProduct.ru_data.isnot(None),
                                     IkeaProduct.ua_data.isnot(None),
                                     cast(IkeaProduct.ua_data, String) != cast(literal(None, JSON()), String),
                                     or_(IkeaProduct.ua_data.isnot(None), IkeaProduct.pl_data.isnot(None))).all()

    for item in items:
        print(item.to_dict().keys())
    groups, groups_lists = get_groups(items)
    df1 = pd.DataFrame.from_dict(groups)

    result = []
    for i in items:
        if i.code.replace(".", "").strip("0") in ignore_codes:
            continue
        if len(get_tags(i)) > 2:
            result.append({
                **get_data(i),
                "Идентификатор_группы": groups_lists.index(get_tags(i)[-2]),
                'Код_товара': i.code
            })

    all_tasks = {i.code: i for i in items}

    for item in result:
        try:
            if item["Код_товара"] in all_tasks:
                available_dict = {
                    True: '+',
                    False: '-',
                    None: ''
                }
                item['Наличие'] = available_dict[all_tasks[item["Код_товара"]].is_available]
            else:
                item['Наличие'] = "-"

            pr = float(re.search(r'\d*[.,]?\d*', item["Цена"]).group(0).replace(',', '.').replace(" ", ""))
            if mode == "UA":
                for p in PRICES:
                    if p > pr:
                        item['Цена'] = pr * PRICES[p]
                        break
            else:
                for p, val in PRICES_PL.items():
                    if p > pr:
                        item['Цена'] = pr * val
                        break
                print(item['Цена'], pr)

            def preparing(cod):
                cod = str(cod)
                # while len(cod) < 8:
                #     cod = "0" + cod
                return f"{cod[:3]}.{cod[3:6]}.{cod[6:]}"

            if item["Код_товара"] in SEARCH_TEXT:
                item["Ключевые_слова"] = SEARCH_TEXT[item["Код_товара"]]
            else:
                item["Ключевые_слова"] = ",".join(filter(lambda x: len(x) > 3 and not re.search("\d{2,}", x),
                                                         item["Название_позиции"].lower().split(" ")))

            item["Идентификатор_товара"] = item['Код_товара']
            item["Код_товара"] = preparing(item['Код_товара'])
            item["Валюта"] = "UAH"
            item["Цена"] = float(str(item["Цена"]).replace(" ", "").replace(",", "."))

            item["Название_позиции"] = item["Название_позиции"][:100]
            item["Ссылка_изображения"] = ", ".join(item["Ссылка_изображения"].split(", ")[:9])
            if "Примечание" in item:
                del item["Примечание"]
            if 'tags' in item:
                del item['tags']
        except:
            logger.exception('Fail to format product')

    products_sheet = f'{PROJECT_DIR}/data/Export Products Sheet.csv'
    groups_sheet = f'{PROJECT_DIR}/data/Export Groups Sheet.csv'
    output = f'{PROJECT_DIR}/data/output.xlsx'
    df1.to_csv(products_sheet, index=False)
    pd.DataFrame.from_dict(result).to_csv(products_sheet, index=False)

    merge_all_to_a_book([products_sheet, groups_sheet], output)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    for sheet in wb:
        sheet_name = sheet.title
        sheet.title = sheet_name.replace(".csv", "")
    wb.save(output)

    return result


def send_mail(email, file, body="", subject="Ikea Bot Data"):
    login = "no-reply@neafiol.site"
    password = "adminnef44!"

    msg = MIMEMultipart(policy=SMTPUTF8)
    msg['From'] = login
    msg['To'] = email
    msg['Subject'] = subject

    part = MIMEBase('application', "octet-stream")
    part.set_payload(file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',
                    'attachment; filename="{}"'.format(op.basename("Data.xlsx")))
    msg.attach(part)

    server = smtplib.SMTP('smtp.yandex.com', 587)
    server.starttls()
    server.login(login, password)
    server.send_message(msg)
    server.quit()
